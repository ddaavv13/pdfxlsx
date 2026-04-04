#!/Library/Developer/CommandLineTools/usr/bin/python3
import io
import os
import re
import uuid
import time
import threading
import statistics
from collections import defaultdict
from typing import Any, Dict, List, Tuple, Optional

import pymupdf as fitz
from PIL import Image
from flask import Flask, request, send_file, render_template_string, abort, jsonify, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from werkzeug.utils import secure_filename

# OCR backends (optional)
try:
    import pytesseract  # type: ignore
except Exception:
    pytesseract = None

try:
    import easyocr  # type: ignore
    import numpy as np
    _easyocr_reader = None  # lazy init
except Exception:
    easyocr = None
    np = None

TESSERACT_EXE = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESSDATA_DIR = r"C:\Program Files\Tesseract-OCR\tessdata"

if pytesseract is not None:
    if os.path.exists(TESSERACT_EXE):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE
    if os.path.isdir(TESSDATA_DIR):
        os.environ["TESSDATA_PREFIX"] = TESSDATA_DIR


def get_easyocr_reader(lang: str = "fr"):
    """Lazy-init EasyOCR reader (heavy model load)."""
    global _easyocr_reader
    if easyocr is None:
        raise RuntimeError("EasyOCR non installé. pip install easyocr numpy")
    if _easyocr_reader is None:
        lang_list = [l.strip() for l in lang.replace("+", ",").split(",") if l.strip()]
        if not lang_list:
            lang_list = ["fr"]
        # Map tesseract lang codes to easyocr codes
        LANG_MAP = {"fra": "fr", "eng": "en", "deu": "de", "spa": "es", "ita": "it"}
        lang_list = [LANG_MAP.get(l, l) for l in lang_list]
        _easyocr_reader = easyocr.Reader(lang_list, gpu=False, verbose=False)
    return _easyocr_reader

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

JOBS: Dict[str, Dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()

HTML_FORM = r"""
<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Convertisseur PDF</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
html,body{margin:0;min-height:100vh;overflow-x:hidden}
body{font-family:system-ui,-apple-system,sans-serif;background:#e8f0f8;display:flex;flex-direction:column;align-items:center}
.header{width:100%;background:#005EA5;color:#fff;padding:20px 0;text-align:center;flex-shrink:0;overflow:hidden;box-sizing:border-box}
.header h1{font-size:22px;font-weight:700;letter-spacing:-0.5px}
.header .sub{font-size:13px;opacity:0.85;margin-top:4px}
.header-inner{text-align:center;padding:8px 16px;position:relative;max-width:900px;margin:0 auto;box-sizing:border-box}
.lang-switch{position:absolute;right:16px;top:50%;transform:translateY(-50%);display:flex;gap:6px}
.logo{position:absolute;left:16px;top:50%;transform:translateY(-50%)}
.flag{font-size:22px;cursor:pointer;opacity:0.5;transition:opacity 0.2s}
.flag:hover{opacity:0.8}
.flag.active{opacity:1}

.card{background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,0.08);max-width:900px;width:95%;margin:24px auto 24px;padding:28px 32px}
.stitle{font-size:15px;font-weight:700;color:#005EA5;margin:20px 0 10px}
.stitle:first-child{margin-top:0}
.drop-zone{border:2px dashed #b0c4de;border-radius:12px;padding:18px 20px;text-align:center;cursor:pointer;transition:all 0.2s;background:#f8fafc}
.drop-zone:hover,.drop-zone.dragover{border-color:#00A3E0;background:#e8f4fd}
.drop-zone .icon{font-size:42px;color:#00A3E0}
.drop-zone p{color:#555;margin-top:8px;font-size:14px}
.drop-zone .fn{color:#005EA5;font-weight:600;font-size:15px;margin-top:8px}
.drop-zone input{display:none}
.lang-row{display:flex;align-items:center;gap:12px;margin-top:12px}
.lang-row label{font-size:14px;font-weight:600;color:#333}
.lang-row select{padding:8px 14px;border:1px solid #ccc;border-radius:8px;font-size:14px;background:#fff}
#prevSec{display:none;margin-top:16px}
.prev-wrap{position:relative;border:1px solid #ddd;border-radius:10px;overflow:hidden;background:#eee;display:flex;justify-content:center}
.prev-wrap canvas{max-width:100%;cursor:crosshair}
.pnav{display:flex;align-items:center;justify-content:center;gap:12px;margin-top:10px}
.pnav button{padding:6px 16px;border:1px solid #ccc;border-radius:8px;background:#fff;cursor:pointer;font-size:13px;font-weight:600}
.pnav button:hover{background:#e8f4fd;border-color:#00A3E0}
.pnav button:disabled{opacity:0.35;cursor:not-allowed;pointer-events:none}
.pnav span{font-size:13px;color:#555}
.hbar{border:1px solid #bfdbfe;transition:background 0.3s;border-radius:8px;padding:8px 12px;margin-bottom:10px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.hbar-icon{flex-shrink:0}
.hbar-cols{display:flex;gap:4px;flex-wrap:wrap}
.hcol{background:#dbeafe;color:#1d4ed8;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.hbar-btn{padding:3px 10px;border:1px solid #93c5fd;border-radius:6px;background:#fff;font-size:11px;cursor:pointer;color:#1d4ed8;margin-left:auto}
.hbar-btn:hover{background:#eff6ff}
.hbar.detecting{background:#fff1f2;border-color:#fecdd3}
.hbar.detected{background:#eff6ff;border-color:#bfdbfe}
.hbar-btn:disabled{opacity:0.35;cursor:not-allowed;pointer-events:none}
.hbar-wait{font-size:12px;font-style:italic}.detecting .hbar-wait{color:#be123c;font-size:13px;font-weight:600;font-style:normal}.detected .hbar-wait{color:#6b7280}
.htab{padding:3px 10px;border:1px solid #93c5fd;border-radius:6px;background:#fff;font-size:11px;cursor:pointer;color:#1d4ed8}
.htab:hover{background:#eff6ff}
.htab.active{background:#dbeafe;border-color:#1d4ed8;font-weight:700}
.cbar{border:1px solid #fed7aa;border-radius:8px;padding:8px 12px;margin-bottom:10px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;background:#fff7ed}
.cbar-btn{padding:3px 10px;border:1px solid #fdba74;border-radius:6px;background:#fff;font-size:11px;cursor:pointer;color:#ea580c}
.cbar-btn:hover{background:#fff7ed}
.cbar-btn.active{background:#fed7aa;border-color:#ea580c;font-weight:700}
.cbar-btn:disabled{opacity:0.35;cursor:not-allowed;pointer-events:none}
.cbar-info{font-size:12px;color:#ea580c;display:flex;gap:4px;flex-wrap:wrap;align-items:center}
.ccol{background:#ffedd5;color:#ea580c;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.zbar{border:1px solid #e2e8f0;border-radius:8px;padding:6px 12px;margin-bottom:10px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;background:#f8fafc;min-height:28px}
.zbar-hint{font-size:12px;color:#94a3b8;font-style:italic}
.clear-btn{padding:3px 10px;border:1px solid #e2e8f0;border-radius:6px;background:#fff;font-size:11px;cursor:pointer;color:#64748b;margin-left:auto}
.clear-btn:hover{background:#fef2f2;border-color:#fca5a5;color:#dc2626}
.clear-btn:disabled{opacity:0.35;cursor:not-allowed;pointer-events:none}
.zone-hint{font-size:12px;color:#999;margin-top:8px}
.zone-list{display:flex;gap:6px;margin-top:6px;flex-wrap:wrap}
.zone-tag{background:#fee2e2;color:#b91c1c;padding:3px 10px;border-radius:6px;font-size:11px;display:inline-flex;align-items:center;gap:4px}
.zone-tag{padding:6px 10px;font-size:12px;gap:6px;align-items:center}
.zone-tag b{color:#333}
.zrange{color:#888}
.zbtn{cursor:pointer;padding:3px 8px;border-radius:5px;font-size:11px;font-weight:600;display:inline-block}
.zback{background:#e0f2fe;color:#005EA5;border:1px solid #93c5fd}
.zback:hover{background:#bfdbfe}
.zfwd{background:#e0f2fe;color:#005EA5;border:1px solid #93c5fd}
.zfwd:hover{background:#bfdbfe}
.zdel{background:#fee2e2;color:#dc2626;border:1px solid #fca5a5}
.zdel:hover{background:#fecaca}
.cbtn{margin-top:20px;padding:14px 32px;border:0;border-radius:10px;background:linear-gradient(135deg,#005EA5,#00A3E0);color:#fff;font-size:16px;font-weight:700;cursor:pointer;width:100%;transition:transform 0.1s}
.cbtn:hover{transform:scale(1.01)}
.cbtn:disabled{opacity:0.5;cursor:not-allowed;transform:none}
.cbtn.cancel{background:linear-gradient(135deg,#dc2626,#ef4444)}
.cbtn.download{background:linear-gradient(135deg,#16a34a,#22c55e)}
#progBox{display:none;margin-top:20px}
.bwrap{width:100%;height:10px;background:#e2e8f0;border-radius:8px;overflow:hidden}
.bfill{width:0%;height:100%;background:linear-gradient(90deg,#005EA5,#00A3E0);transition:width 0.4s;border-radius:8px}
.stxt{margin-top:10px;font-size:14px;color:#333}
.stxt.err{color:#dc2626;font-weight:600}
.stxt.ok{color:#16a34a;font-weight:600}
.ft{padding:12px;text-align:center;font-size:11px;color:#8facc4}
.ft{padding:16px;text-align:center;font-size:11px;color:#aaa}
</style>
</head>
<body>
<div class="header">
<div class="header-inner">
<div class="logo"><img src="/logo.png" alt="Logo" style="height:40px;opacity:0.85"></div>
<div class="title-wrap"><h1 id="title">Convertisseur PDF</h1></div>
<div class="lang-switch"><span class="flag active" onclick="setLang('fr')" id="flagFr" title="Français">&#127467;&#127479;</span><span class="flag" onclick="setLang('en')" id="flagEn" title="English">&#127468;&#127463;</span></div>
</div>
<div class="card">
<div class="stitle">&#128196; <span id="t_sel">Sélectionnez votre fichier PDF</span></div>
<div class="drop-zone" id="dz" onclick="document.getElementById('fi').click()">
<div class="icon">&#128206;</div>
<p><span id="t_drop">Glissez votre PDF ici ou <b>cliquez pour parcourir</b></span></p>
<div class="fn" id="fn"></div>
<input type="file" id="fi" accept=".pdf,application/pdf">
</div>
<div class="lang-row" id="langRow" style="display:none">
<label id="t_lang">Langue du document :</label>
<select id="lg">
<option value="fr" selected id="opt_fr">Français</option>
<option value="en" id="opt_en">Anglais</option>
</select>
<span style="font-size:11px;color:#999;margin-left:4px" id="t_langhelp">pour la reconnaissance du texte</span>
</div>
<div id="prevSec">
<div id="headerBar" class="hbar" style="display:none">
<svg class="hbar-icon" width="18" height="14" viewBox="0 0 18 14"><rect x="0" y="0" width="18" height="4" rx="1" fill="#2563eb"/><rect x="0" y="5" width="18" height="2" rx="0.5" fill="#2563eb" opacity="0.3"/><rect x="0" y="8.5" width="18" height="2" rx="0.5" fill="#2563eb" opacity="0.3"/><rect x="0" y="12" width="18" height="2" rx="0.5" fill="#2563eb" opacity="0.3"/></svg>
<span id="hdrTabs" style="display:flex;gap:4px"></span>
<span class="hbar-cols" id="hbarCols"></span>
<div id="hbarBtns" style="margin-left:auto;display:flex;gap:4px"><button class="hbar-btn" onclick="redetectHeader()" id="t_hsauto">Auto</button><button class="hbar-btn" onclick="moveHeader()" id="t_hsmove">Déplacer</button><button class="hbar-btn" onclick="addHeader()" id="t_hsadd" title="Ajouter un en-tête">+</button><button class="hbar-btn" onclick="removeHeader()" id="t_hsrm" disabled title="Supprimer cet en-tête">&minus;</button></div>
</div>
<div id="colBar" class="cbar" style="display:none">
<svg width="18" height="14" viewBox="0 0 18 14"><rect x="0" y="0" width="4" height="14" rx="1" fill="#ea580c" opacity="0.3"/><rect x="7" y="0" width="4" height="14" rx="1" fill="#ea580c" opacity="0.3"/><rect x="14" y="0" width="4" height="14" rx="1" fill="#ea580c" opacity="0.3"/></svg>
<span class="cbar-info" id="colBarInfo"></span>
<div id="colBarBtns" style="margin-left:auto;display:flex;gap:4px"><button class="cbar-btn" onclick="autoColLines()" id="t_colauto">Auto</button><button class="cbar-btn" onclick="moveColLines()" id="t_colmove2">Déplacer</button><button class="cbar-btn" onclick="addColLine()" id="t_coladd" title="Ajouter une ligne">+</button><button class="cbar-btn" onclick="removeColLineMode()" id="t_colrm" title="Supprimer une ligne">&minus;</button></div>
</div>
<div id="zoneBar" class="zbar">
<span class="zbar-hint" id="zh">Dessinez un rectangle sur les zones à ignorer</span>
<span id="zl" style="display:flex;gap:4px;flex-wrap:wrap"></span>
<button class="clear-btn" onclick="clearZones()" id="t_clear" style="display:none">&#128465;</button>
</div>
<div class="stitle">&#128065; <span id="t_prev">Aperçu</span></div>
<div class="prev-wrap"><canvas id="cv"></canvas></div>
<div class="pnav">
<button onclick="cp(-1)">&laquo;</button>
<span id="pi">Page 1/1</span>
<button onclick="cp(1)">&raquo;</button>
</div>

</div>
<button class="cbtn" id="cb" onclick="mainAction()" disabled style="display:none">&#9989; Convertir en Excel</button>
<div id="progBox">
<div class="bwrap"><div class="bfill" id="bf"></div></div>
<div class="stxt" id="st"></div>
</div>
</div>

<script>
let fid=null,tp=0,pg=1,pimg=null,currentJobId=null,currentTimer=null;
const dz=document.getElementById('dz'),fi=document.getElementById('fi'),cv=document.getElementById('cv'),cx=cv.getContext('2d');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('dragover')});
dz.addEventListener('dragleave',()=>dz.classList.remove('dragover'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('dragover');if(e.dataTransfer.files.length){fi.files=e.dataTransfer.files;hf()}});
fi.addEventListener('change',hf);
async function hf(){
const f=fi.files[0];if(!f)return;
document.getElementById('fn').textContent=f.name;
document.getElementById('cb').disabled=true;
const fd=new FormData();fd.append('pdf',f);
try{const r=await fetch('/upload',{method:'POST',body:fd});const d=await r.json();
if(d.error){alert(d.error);return}
fid=d.file_id;tp=d.total_pages;pg=1;ez=[];headers=[];activeHdr=0;exitColMode();document.getElementById('colBar').style.display='none';
document.getElementById('prevSec').style.display='block';
document.getElementById('langRow').style.display='';
document.getElementById('cb').style.display='';document.getElementById('cb').disabled=false;
document.getElementById('progBox').style.display='none';setBtnState('convert');
lp();
}catch(e){alert((uiLang==='fr'?'Erreur: ':'Error: ')+e.message)}}
async function lp(){
const img=new Image();
img.onload=()=>{pimg=img;cv.width=img.width;cv.height=img.height;rd()};
img.src='/preview/'+fid+'/'+pg+'?t='+Date.now();
document.getElementById('pi').textContent='Page '+pg+'/'+tp;
if(pg===1&&headers.length===0){autoDetectHeader()}}
async function autoDetectHeader(){
document.getElementById('headerBar').style.display='flex';document.getElementById('headerBar').className='hbar detecting';
document.getElementById('t_hsmove').disabled=true;document.getElementById('t_hsauto').disabled=true;
document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Détection des en-têtes de colonnes...':'Detecting column headers...')+'</span>';lockUI(true);
try{const r=await fetch('/auto_detect_header/'+fid+'/1');const d=await r.json();
if(d.columns&&d.columns.length>=3){
const pw=d.page_width||0;
const h={page:pg,lineY:d.y,cols:d.columns,pos:d.positions,colLines:pw>0?d.positions.map(x=>x/pw):[],pageWidth:pw};
if(headers.length===0){headers.push(h);activeHdr=0}else{headers[activeHdr]=h}
lockUI(false);showHeaderResult()}
else{document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Non détecté - cliquez sur la ligne d\'en-tête':'Not detected - click on the header row')+'</span>';lockUI(false);}
}catch(e){document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">Erreur</span>'}}
function showHeaderResult(){
document.getElementById('headerBar').style.display='flex';document.getElementById('headerBar').className='hbar detected';
document.getElementById('t_hsmove').disabled=false;document.getElementById('t_hsauto').disabled=false;
document.getElementById('t_hsrm').disabled=headers.length<=1;
const h=curHdr();
document.getElementById('hbarCols').innerHTML=h?h.cols.map(c=>'<span class="hcol">'+c+'</span>').join(''):'';
updateHdrTabs();
if(h&&h.colLines.length>0)updateColBar();
rd()}
function moveHeader(){
headerMode=false;movingHeader=true;exitColMode();lockUI(true);
document.getElementById('cv').style.pointerEvents='auto';
document.getElementById('cv').style.opacity='1';
document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Glissez la ligne bleue vers l\'en-tête':'Drag the blue line to the header')+'</span>';
cv.style.cursor='ns-resize';
cv.onmousedown=startDragHeader;cv.onmousemove=dragHeader;cv.onmouseup=endDragHeader;
}
let draggingHeader=false;let movingHeader=false;let addingHeader=false;
function startDragHeader(e){draggingHeader=true}
function dragHeader(e){
if(!pimg)return;
const r=cv.getBoundingClientRect();
const y=(e.clientY-r.top)*(cv.height/r.height);
if(draggingHeader){if(curHdr())curHdr().lineY=y/cv.height;previewHeaderY=null}
else{previewHeaderY=y/cv.height}
rd()}
function endDragHeader(e){
if(!draggingHeader)return;
draggingHeader=false;dr=false;ds=null;
cv.style.cursor='crosshair';
cv.onmousedown=null;cv.onmousemove=null;cv.onmouseup=null;
// Re-detect columns at new position
lockUI(true);
document.getElementById('headerBar').className='hbar detecting';
document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Détection...':'Detecting...')+'</span>';
clickHeader(curHdr()?curHdr().lineY:0).then(()=>{
// Restore normal canvas handlers
restoreCanvasHandlers()})}
function restoreCanvasHandlers(){movingHeader=false;addingHeader=false;previewHeaderY=null;dr=false;ds=null;
cv.onmousedown=null;cv.onmousemove=null;cv.onmouseup=null;
cv.style.cursor='crosshair';
lockUI(false);
}
function redetectHeader(){
if(curHdr()){curHdr().cols=[];curHdr().pos=[];curHdr().lineY=null;curHdr().colLines=[]}
exitColMode();document.getElementById('colBar').style.display='none';
document.getElementById('headerBar').className='hbar detecting';lockUI(true);autoDetectHeader()}
function cp(d){const p=pg+d;if(p<1||p>tp)return;pg=p;lp()}
let ez=[];let dr=false;let ds=null;
let uiLocked=false;
let headers=[];let activeHdr=0;
let colEditMode=null;let draggingColIdx=-1;let previewColX=null;let previewHeaderY=null;
function curHdr(){return headers[activeHdr]||null}
const ZCOLORS=[
{fill:'rgba(220,38,38,0.20)',stroke:'#dc2626',tag:'#fee2e2',text:'#b91c1c'},
{fill:'rgba(37,99,235,0.20)',stroke:'#2563eb',tag:'#dbeafe',text:'#1d4ed8'},
{fill:'rgba(22,163,74,0.20)',stroke:'#16a34a',tag:'#dcfce7',text:'#15803d'},
{fill:'rgba(234,88,12,0.20)',stroke:'#ea580c',tag:'#ffedd5',text:'#c2410c'},
{fill:'rgba(147,51,234,0.20)',stroke:'#9333ea',tag:'#f3e8ff',text:'#7e22ce'},
{fill:'rgba(6,182,212,0.20)',stroke:'#06b6d4',tag:'#cffafe',text:'#0891b2'},
{fill:'rgba(236,72,153,0.20)',stroke:'#ec4899',tag:'#fce7f3',text:'#be185d'},
{fill:'rgba(202,138,4,0.20)',stroke:'#ca8a04',tag:'#fef9c3',text:'#a16207'},
{fill:'rgba(79,70,229,0.20)',stroke:'#4f46e5',tag:'#e0e7ff',text:'#4338ca'},
{fill:'rgba(5,150,105,0.20)',stroke:'#059669',tag:'#d1fae5',text:'#047857'},
{fill:'rgba(239,68,68,0.20)',stroke:'#ef4444',tag:'#fee2e2',text:'#dc2626'},
{fill:'rgba(245,158,11,0.20)',stroke:'#f59e0b',tag:'#fef3c7',text:'#d97706'},
{fill:'rgba(168,85,247,0.20)',stroke:'#a855f7',tag:'#ede9fe',text:'#7c3aed'},
{fill:'rgba(20,184,166,0.20)',stroke:'#14b8a6',tag:'#ccfbf1',text:'#0d9488'},
{fill:'rgba(244,63,94,0.20)',stroke:'#f43f5e',tag:'#ffe4e6',text:'#e11d48'},
{fill:'rgba(132,204,22,0.20)',stroke:'#84cc16',tag:'#ecfccb',text:'#65a30d'},
{fill:'rgba(99,102,241,0.20)',stroke:'#6366f1',tag:'#e0e7ff',text:'#4f46e5'},
{fill:'rgba(249,115,22,0.20)',stroke:'#f97316',tag:'#fff7ed',text:'#ea580c'},
{fill:'rgba(34,197,94,0.20)',stroke:'#22c55e',tag:'#dcfce7',text:'#16a34a'},
{fill:'rgba(217,70,239,0.20)',stroke:'#d946ef',tag:'#fae8ff',text:'#c026d3'},
{fill:'rgba(14,165,233,0.20)',stroke:'#0ea5e9',tag:'#e0f2fe',text:'#0284c7'},
{fill:'rgba(251,146,60,0.20)',stroke:'#fb923c',tag:'#ffedd5',text:'#ea580c'},
{fill:'rgba(52,211,153,0.20)',stroke:'#34d399',tag:'#d1fae5',text:'#059669'},
{fill:'rgba(192,38,211,0.20)',stroke:'#c026d3',tag:'#fae8ff',text:'#a21caf'},
{fill:'rgba(56,189,248,0.20)',stroke:'#38bdf8',tag:'#e0f2fe',text:'#0284c7'},
{fill:'rgba(163,230,53,0.20)',stroke:'#a3e635',tag:'#ecfccb',text:'#65a30d'},
{fill:'rgba(251,113,133,0.20)',stroke:'#fb7185',tag:'#ffe4e6',text:'#e11d48'},
{fill:'rgba(45,212,191,0.20)',stroke:'#2dd4bf',tag:'#ccfbf1',text:'#0d9488'},
{fill:'rgba(253,186,116,0.20)',stroke:'#fdba74',tag:'#fff7ed',text:'#c2410c'},
{fill:'rgba(129,140,248,0.20)',stroke:'#818cf8',tag:'#e0e7ff',text:'#4338ca'},
{fill:'rgba(74,222,128,0.20)',stroke:'#4ade80',tag:'#dcfce7',text:'#15803d'},
{fill:'rgba(232,121,249,0.20)',stroke:'#e879f9',tag:'#fae8ff',text:'#a21caf'},
];
function zcolor(i){return ZCOLORS[i%ZCOLORS.length]}
function rd(){
if(!pimg)return;cx.drawImage(pimg,0,0);
for(let hi=0;hi<headers.length;hi++){const hh=headers[hi];if(hh.lineY!==null&&pg===hh.page){const hy=hh.lineY*cv.height+cv.height*0.008;const isActive=hi===activeHdr;cx.strokeStyle=isActive?'#2563eb':'rgba(100,116,139,0.4)';cx.lineWidth=isActive?2:1;cx.setLineDash([6,4]);cx.beginPath();cx.moveTo(0,hy);cx.lineTo(cv.width,hy);cx.stroke();cx.setLineDash([]);cx.fillStyle=isActive?'#2563eb':'rgba(100,116,139,0.5)';cx.font='bold '+(isActive?'14':'11')+'px system-ui';const htxt=(uiLang==='fr'?'En-tête ':'Header ')+(hi+1);cx.fillText(htxt,cv.width-cx.measureText(htxt).width-8,hy-3)}}
for(const z of ez){
if(z.a||(pg>=z.fromPage&&pg<=z.toPage)){
const rx=z.x*cv.width,ry=z.y*cv.height,rw=z.w*cv.width,rh=z.h*cv.height;
const zc=zcolor(ez.indexOf(z));cx.fillStyle=zc.fill;cx.fillRect(rx,ry,rw,rh);
cx.strokeStyle=zc.stroke;cx.lineWidth=1.5;cx.strokeRect(rx,ry,rw,rh);
cx.fillStyle=zc.stroke;cx.font='11px system-ui';
cx.fillText(z.fromPage===z.toPage?'Page '+z.fromPage:(uiLang==='fr'?'Pages ':'Pages ')+z.fromPage+'-'+z.toPage,rx+4,ry+13);
cx.font='bold 15px system-ui';cx.fillText('\u00d7',rx+rw-14,ry+14)}}
{const h=curHdr();const cl=h?h.colLines:[];
if(cl.length>0&&pg===h.page){for(let ci=1;ci<cl.length;ci++){const clx=cl[ci]*cv.width;cx.strokeStyle=draggingColIdx===ci?'rgba(234,88,12,0.8)':'rgba(234,88,12,0.45)';cx.lineWidth=draggingColIdx===ci?2.5:1.5;cx.setLineDash([4,4]);cx.beginPath();cx.moveTo(clx,0);cx.lineTo(clx,cv.height);cx.stroke();cx.setLineDash([])}}}
if(previewColX!==null){const px=previewColX*cv.width;cx.strokeStyle='rgba(234,88,12,0.5)';cx.lineWidth=2;cx.setLineDash([4,4]);cx.beginPath();cx.moveTo(px,0);cx.lineTo(px,cv.height);cx.stroke();cx.setLineDash([])}
if(previewHeaderY!==null){const py=previewHeaderY*cv.height;cx.strokeStyle='rgba(37,99,235,0.5)';cx.lineWidth=2;cx.setLineDash([4,4]);cx.beginPath();cx.moveTo(0,py);cx.lineTo(cv.width,py);cx.stroke();cx.setLineDash([])}
updateZL()}
function updateZL(){
const el=document.getElementById('zl');
el.innerHTML=ez.map((z,i)=>{
const f=uiLang==='fr';
const range=z.fromPage===z.toPage?'p.'+z.fromPage:'p.'+z.fromPage+'-'+z.toPage;
const canBack=z.fromPage>1;
const canFwd=(z.toPage||z.fromPage)<tp;
const zc=zcolor(i);
let h='<span class="zone-tag" style="background:'+zc.tag+';color:'+zc.text+'">';
if(canBack)h+='<span class="zbtn zback" onclick="extendBack('+i+')" title="'+(f?'Étendre':'Extend')+'">&#9664;</span> ';
h+='<b>Z'+(i+1)+'</b> ('+range+') ';
if(canFwd)h+='<span class="zbtn zfwd" onclick="extendZone('+i+')" title="'+(f?'Étendre':'Extend')+'">&#9654;</span> ';
h+='<span class="zbtn zdel" onclick="rmz('+i+')" title="'+(f?'Supprimer':'Delete')+'">&#10005;</span></span>';
return h}).join('');
const zh=document.getElementById('zh');
const cl=document.getElementById('t_clear');
if(ez.length){zh.style.display='none';cl.style.display=''}
else{zh.style.display='';cl.style.display='none';zh.textContent=uiLang==='fr'?'Dessinez un rectangle sur les zones à ignorer':'Draw a rectangle on zones to exclude'}}
function extendZone(i){ez[i].toPage=tp;ez[i].a=false;rd()}
function extendBack(i){ez[i].fromPage=1;ez[i].a=false;rd()}
function rmz(i){ez.splice(i,1);rd()}
function clearZones(){ez=[];rd()}
cv.addEventListener('mousedown',e=>{
if(addingHeader||movingHeader)return;
if(colEditMode){handleColDown(e);return}
const r=cv.getBoundingClientRect();
const sx=(e.clientX-r.left)*(cv.width/r.width);
const sy=(e.clientY-r.top)*(cv.height/r.height);
for(let i=ez.length-1;i>=0;i--){const z=ez[i];
if(!(z.a||(pg>=z.fromPage&&pg<=z.toPage)))continue;
const rx=z.x*cv.width,ry=z.y*cv.height,rw=z.w*cv.width;
if(sx>=rx+rw-20&&sx<=rx+rw+2&&sy>=ry&&sy<=ry+20){
if(z.fromPage===z.toPage){ez.splice(i,1)}
else if(pg===z.fromPage){z.fromPage++}
else if(pg===z.toPage){z.toPage--}
else{const z2={p:pg,x:z.x,y:z.y,w:z.w,h:z.h,a:false,fromPage:pg+1,toPage:z.toPage};z.toPage=pg-1;ez.push(z2)}
rd();return}}
dr=true;ds={x:sx/cv.width,y:sy/cv.height}});
cv.addEventListener('mousemove',e=>{
if(addingHeader||movingHeader)return;
if(colEditMode){handleColMove(e);return}
if(!dr||!ds||uiLocked)return;
const r=cv.getBoundingClientRect();
const mx=(e.clientX-r.left)*(cv.width/r.width)/cv.width;
const my=(e.clientY-r.top)*(cv.height/r.height)/cv.height;
rd();
const rx=ds.x*cv.width,ry=ds.y*cv.height;
const rw=(mx-ds.x)*cv.width,rh=(my-ds.y)*cv.height;
const nc=zcolor(ez.length);cx.fillStyle=nc.fill;cx.fillRect(rx,ry,rw,rh);
cx.strokeStyle=nc.stroke;cx.lineWidth=1.5;cx.setLineDash([5,3]);cx.strokeRect(rx,ry,rw,rh);cx.setLineDash([])});
cv.addEventListener('mouseup',e=>{
if(addingHeader||movingHeader)return;
if(colEditMode){handleColUp(e);return}
if(!dr||!ds)return;dr=false;
const r=cv.getBoundingClientRect();
const ex2=(e.clientX-r.left)*(cv.width/r.width)/cv.width;
const ey2=(e.clientY-r.top)*(cv.height/r.height)/cv.height;
const x=Math.min(ds.x,ex2),y=Math.min(ds.y,ey2);
const w=Math.abs(ex2-ds.x),h=Math.abs(ey2-ds.y);
ds=null;
if(w>0.01&&h>0.01){
ez.push({p:pg,x,y,w,h,a:false,fromPage:pg,toPage:pg})}
rd()})

let btnState='convert';let dlUrl='';
function setBtnState(state,url){
btnState=state;const cb=document.getElementById('cb');
if(state==='convert'){cb.className='cbtn';cb.disabled=false;cb.innerHTML=(uiLang==='fr'?'&#9989; Convertir en Excel':'&#9989; Convert to Excel')}
else if(state==='cancel'){cb.className='cbtn cancel';cb.disabled=false;cb.innerHTML=(uiLang==='fr'?'&#10060; Annuler':'&#10060; Cancel')}
else if(state==='download'){cb.className='cbtn download';cb.disabled=false;cb.innerHTML=(uiLang==='fr'?'&#128229; Télécharger le fichier Excel':'&#128229; Download Excel file');dlUrl=url||''}
}
function mainAction(){
if(btnState==='convert')go();
else if(btnState==='cancel')cancelJob();
else if(btnState==='download'){window.location.href=dlUrl;setBtnState('convert')}
}
async function cancelJob(){
if(!currentJobId)return;
try{await fetch('/cancel/'+currentJobId,{method:'POST'})}catch(e){}
if(currentTimer)clearInterval(currentTimer);
currentTimer=null;currentJobId=null;
document.getElementById('st').textContent=uiLang==='fr'?'Conversion annulée.':'Conversion cancelled.';
document.getElementById('st').className='stxt err';
document.getElementById('bf').style.width='0%';
setBtnState('convert');lockUI(false)}
async function go(){
const cb=document.getElementById('cb'),bf=document.getElementById('bf'),st=document.getElementById('st'),pb=document.getElementById('progBox');
lockUI(true);pb.style.display='block';bf.style.width='0%';st.textContent=uiLang==='fr'?'Envoi...':'Sending...';st.className='stxt';
setBtnState('cancel');
try{const r=await fetch('/start',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({file_id:fid,lang:document.getElementById('lg').value,exclude_zones:ez,headers:headers.map(h=>({page:h.page,lineY:h.lineY,cols:h.cols,positions:h.pageWidth>0?h.colLines.map(x=>x*h.pageWidth):h.pos}))})});
const d=await r.json();if(d.error){st.textContent=d.error;st.className='stxt err';setBtnState('convert');lockUI(false);return}
const jid=d.job_id;currentJobId=jid;
currentTimer=setInterval(async()=>{try{const s=await(await fetch('/status/'+jid)).json();
bf.style.width=(s.progress||0)+'%';st.textContent=s.message||'';st.className='stxt';
if(s.status==='done'){clearInterval(currentTimer);currentTimer=null;bf.style.width='100%';st.textContent=uiLang==='fr'?'Conversion terminée !':'Conversion complete!';currentJobId=null;st.className='stxt ok';setBtnState('download','/download/'+jid);lockUI(false)}
else if(s.status==='error'){clearInterval(currentTimer);currentTimer=null;st.textContent=s.error||(uiLang==='fr'?'Erreur':'Error');st.className='stxt err';setBtnState('convert');lockUI(false);currentJobId=null}
}catch(e){clearInterval(currentTimer);currentTimer=null;st.textContent=uiLang==='fr'?'Connexion perdue':'Connection lost';st.className='stxt err';setBtnState('convert');lockUI(false)}},2000);
}catch(e){st.textContent=(uiLang==='fr'?'Erreur: ':'Error: ')+e.message;st.className='stxt err';setBtnState('convert');lockUI(false)}}

const TR={
fr:{title:'Convertisseur PDF → XLSX',sel:'Sélectionnez votre fichier PDF',drop:'Glissez votre PDF ici ou <b>cliquez pour parcourir</b>',lang:'Langue du document :',langhelp:'pour la reconnaissance du texte',prev:'Aperçu',conv:'\u2705 Convertir en Excel',cancel:'Annuler',dl:'\ud83d\udce5 Télécharger le fichier Excel',clear:'\ud83d\uddd1 Effacer les zones',hint:'Dessinez un rectangle sur les zones à ignorer',zones:' zone(s) active(s) sur cette page',prep:'Préparation...',colauto:'Auto',colmove:'Déplacer',},
en:{title:'PDF → XLSX Converter',sel:'Select your PDF file',drop:'Drag your PDF here or <b>click to browse</b>',lang:'Document language:',langhelp:'for text recognition',prev:'Preview',conv:'\u2705 Convert to Excel',cancel:'Cancel',dl:'\ud83d\udce5 Download Excel file',clear:'\ud83d\uddd1 Clear zones',hint:'Draw a rectangle on zones to exclude',zones:' active zone(s) on this page',prep:'Preparing...',colauto:'Auto',colmove:'Move',}
};
let uiLang=localStorage.getItem('pdfxlsx_lang')||'fr';
function setLang(l){
uiLang=l;localStorage.setItem('pdfxlsx_lang',l);
document.getElementById('flagFr').className='flag'+(l==='fr'?' active':'');
document.getElementById('flagEn').className='flag'+(l==='en'?' active':'');
const t=TR[l];
document.getElementById('title').textContent=t.title;
const s=document.getElementById('t_sel');if(s)s.textContent=t.sel;
const d=document.getElementById('t_drop');if(d)d.innerHTML=t.drop;
const lg=document.getElementById('t_lang');if(lg)lg.textContent=t.lang;
const lh=document.getElementById('t_langhelp');if(lh)lh.textContent=t.langhelp;
const p=document.getElementById('t_prev');if(p)p.textContent=t.prev;
const cv=document.getElementById('cb');if(cv)cv.innerHTML=t.conv;
setBtnState(btnState,dlUrl);
const zh=document.getElementById('zh');if(zh&&!ez.length)zh.textContent=t.hint;
if(typeof rd==='function'&&typeof ez!=='undefined')try{rd()}catch(e){}
const of=document.getElementById('opt_fr');if(of)of.textContent=l==='fr'?'Français':'French';
const oe=document.getElementById('opt_en');if(oe)oe.textContent=l==='fr'?'Anglais':'English';
const ca2=document.getElementById('t_colauto');if(ca2)ca2.textContent=t.colauto;
const cm2=document.getElementById('t_colmove2');if(cm2)cm2.textContent=t.colmove;
if(curHdr()&&curHdr().colLines.length>0)updateColBar();
updateHdrTabs();
}
setLang(uiLang);

async function clickHeader(relY){
try{const r=await fetch('/detect_header_at/'+fid+'/'+pg,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({y:relY})});
const d=await r.json();
if(d.error==='data_row'){
if(curHdr()&&!curHdr().cols.length){headers.splice(activeHdr,1);if(activeHdr>=headers.length)activeHdr=Math.max(0,headers.length-1)}
restoreCanvasHandlers();
if(headers.length>0){showHeaderResult()}else{document.getElementById('headerBar').className='hbar detected';document.getElementById('hbarCols').innerHTML=''}
rd()}
else if(d.columns&&d.columns.length>=3){
const pw=d.page_width||0;
if(curHdr()){curHdr().cols=d.columns;curHdr().pos=d.positions;curHdr().lineY=relY;curHdr().colLines=pw>0?d.positions.map(x=>x/pw):[];curHdr().pageWidth=pw;curHdr().page=pg}
lockUI(false);showHeaderResult()}
else{showHeaderResult()}}
catch(e){alert('Error: '+e.message)}}

function lockUI(lock){
uiLocked=lock;
document.getElementById('cb').disabled=lock;
document.getElementById('t_clear').disabled=lock;
document.getElementById('cv').style.pointerEvents=lock?'none':'auto';
document.getElementById('cv').style.opacity=lock?'0.6':'1';
document.getElementById('lg').disabled=lock;
document.querySelectorAll('.pnav button').forEach(b=>b.disabled=lock);
document.querySelectorAll('.flag').forEach(f=>f.style.pointerEvents=lock?'none':'auto');
document.getElementById('dz').style.pointerEvents=lock?'none':'auto';
document.getElementById('dz').style.opacity=lock?'0.6':'1';
document.getElementById('fi').disabled=lock;
var hm=document.getElementById('t_hsmove');if(hm)hm.disabled=lock;
var ha=document.getElementById('t_hsauto');if(ha)ha.disabled=lock;
var had=document.getElementById('t_hsadd');if(had)had.disabled=lock;
var hrm=document.getElementById('t_hsrm');if(hrm)hrm.disabled=lock||headers.length<=1;
document.querySelectorAll('.cbar-btn').forEach(b=>b.disabled=lock);
document.getElementById('zoneBar').style.pointerEvents=lock?'none':'auto';
document.getElementById('zoneBar').style.opacity=lock?'0.6':'1';
}
function updateColBar(){
const h=curHdr();if(!h)return;
document.getElementById('colBar').style.display='flex';
const cl=h.colLines;const n=cl.length;
let info=n+' '+(uiLang==='fr'?'colonnes':'columns');
if(colEditMode==='move')info=uiLang==='fr'?'Glissez les lignes':'Drag the lines';
else if(colEditMode==='add')info=uiLang==='fr'?'Cliquez pour ajouter une ligne':'Click to add a line';
else if(colEditMode==='remove')info=uiLang==='fr'?'Cliquez sur une ligne à supprimer':'Click a line to remove';
document.getElementById('colBarInfo').innerHTML=colEditMode?'<span style="font-style:italic">'+info+'</span>':h.cols.map(c=>'<span class="ccol">'+c+'</span>').join('')+' <span style="font-size:11px;color:#6b7280">('+n+')</span>';
document.querySelectorAll('.cbar-btn').forEach(b=>b.classList.remove('active'));
if(colEditMode==='move')document.getElementById('t_colmove2').classList.add('active');
if(colEditMode==='add')document.getElementById('t_coladd').classList.add('active');
if(colEditMode==='remove')document.getElementById('t_colrm').classList.add('active');
}
function exitColMode(){colEditMode=null;draggingColIdx=-1;previewColX=null;cv.style.cursor='crosshair';updateColBar();rd()}
function autoColLines(){
exitColMode();const h=curHdr();
if(h&&h.pos.length>0&&h.pageWidth>0){h.colLines=h.pos.map(x=>x/h.pageWidth)}
updateColBar();rd()}
function moveColLines(){
if(colEditMode==='move'){exitColMode();return}
colEditMode='move';cv.style.cursor='col-resize';updateColBar()}
function addColLine(){
if(colEditMode==='add'){exitColMode();return}
colEditMode='add';cv.style.cursor='copy';updateColBar()}
function removeColLineMode(){
if(colEditMode==='remove'){exitColMode();return}
colEditMode='remove';cv.style.cursor='pointer';updateColBar()}
function findNearCol(e){
const h=curHdr();if(!h)return -1;const cl=h.colLines;
const r=cv.getBoundingClientRect();
const mx=(e.clientX-r.left)*(cv.width/r.width);
let best=-1,bestD=12;
for(let i=1;i<cl.length;i++){const d=Math.abs(mx-cl[i]*cv.width);if(d<bestD){bestD=d;best=i}}
return best}
function handleColDown(e){
if(movingHeader)return;const h=curHdr();if(!h)return;
if(colEditMode==='move'){const idx=findNearCol(e);if(idx>=0){draggingColIdx=idx;rd()}}
else if(colEditMode==='add'){
const r=cv.getBoundingClientRect();const mx=(e.clientX-r.left)*(cv.width/r.width)/cv.width;
let ins=h.colLines.length;
for(let i=0;i<h.colLines.length;i++){if(mx<h.colLines[i]){ins=i;break}}
h.colLines.splice(ins,0,mx);h.cols.splice(ins,0,'Col '+(ins+1));
exitColMode();updateColBar();rd()}
else if(colEditMode==='remove'){const idx=findNearCol(e);if(idx>=0&&h.colLines.length>2){h.colLines.splice(idx,1);h.cols.splice(idx,1);exitColMode();updateColBar();rd()}}}
function handleColMove(e){
const h=curHdr();if(!h)return;const cl=h.colLines;
if(colEditMode==='move'&&draggingColIdx>=0){
const r=cv.getBoundingClientRect();const mx=(e.clientX-r.left)*(cv.width/r.width)/cv.width;
const minX=draggingColIdx>0?cl[draggingColIdx-1]+0.01:0.005;
const maxX=draggingColIdx<cl.length-1?cl[draggingColIdx+1]-0.01:0.995;
cl[draggingColIdx]=Math.max(minX,Math.min(maxX,mx));rd()}
else if(colEditMode==='move'){const idx=findNearCol(e);cv.style.cursor=idx>=0?'col-resize':'default'}
else if(colEditMode==='add'){const r=cv.getBoundingClientRect();previewColX=(e.clientX-r.left)*(cv.width/r.width)/cv.width;rd()}
else if(colEditMode==='remove'){const idx=findNearCol(e);cv.style.cursor=idx>=0&&cl.length>2?'pointer':'not-allowed'}}
function handleColUp(e){if(draggingColIdx>=0){draggingColIdx=-1;exitColMode();updateColBar();rd()}}
function updateHdrTabs(){
const el=document.getElementById('hdrTabs');
el.innerHTML=headers.map((h,i)=>'<span class="htab'+(i===activeHdr?' active':'')+'" onclick="switchHdr('+i+')">'+(uiLang==='fr'?'En-tête ':'Header ')+(i+1)+'</span>').join('');
if(headers.length<=1){el.innerHTML='';el.style.display='none'}else{el.style.display='flex'}}
function switchHdr(i){activeHdr=i;exitColMode();showHeaderResult()}
function addHeader(){
headerMode=false;movingHeader=false;addingHeader=true;exitColMode();
lockUI(true);document.getElementById('cv').style.pointerEvents='auto';document.getElementById('cv').style.opacity='1';
document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Cliquez sur la ligne d\'en-tête du nouveau tableau':'Click on the header row of the new table')+'</span>';
cv.style.cursor='crosshair';
cv.onmousemove=function(e){
const r=cv.getBoundingClientRect();previewHeaderY=(e.clientY-r.top)*(cv.height/r.height)/cv.height;rd()};
cv.onmousedown=function(e){
const r=cv.getBoundingClientRect();const relY=(e.clientY-r.top)*(cv.height/r.height)/cv.height;
previewHeaderY=null;addingHeader=false;cv.onmousedown=null;cv.onmousemove=null;cv.style.cursor='crosshair';
const newH={page:pg,lineY:relY,cols:[],pos:[],colLines:[],pageWidth:0};
headers.push(newH);activeHdr=headers.length-1;
lockUI(true);document.getElementById('headerBar').className='hbar detecting';
document.getElementById('hbarCols').innerHTML='<span class="hbar-wait">'+(uiLang==='fr'?'Détection...':'Detecting...')+'</span>';
clickHeader(relY).then(()=>{restoreCanvasHandlers()})}}
function removeHeader(){
if(headers.length<=1)return;
headers.splice(activeHdr,1);
if(activeHdr>=headers.length)activeHdr=headers.length-1;
exitColMode();showHeaderResult()}
</script>
</body></html>
"""

INVALID_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F\ud800-\udfff\ufffe\uffff]")
DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
AMOUNT_RE = re.compile(r"(?<!\d)(?:\d{1,3}(?:[ .]\d{3})*|\d+),\d{2}(?!\d)")

def set_job(job_id: str, **kwargs: Any) -> None:
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(kwargs)

def safe_excel_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\u00a0": " ", "\u2007": " ", "\u202f": " ", "\u200b": "", "\u200c": "", "\u200d": "", "\ufeff": "",
        "’": "'", "‘": "'", "“": '"', "”": '"', "–": "-", "—": "-", "−": "-", "…": "...",
        "ﬁ": "fi", "ﬂ": "fl", "ﬀ": "ff", "ﬃ": "ffi", "ﬄ": "ffl",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = INVALID_XML_RE.sub("", text)
    text = re.sub(r"[ \t]+", " ", text).strip()
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    if len(text) > 32767:
        text = text[:32767]
    return text

def has_broken_fonts(page: fitz.Page) -> bool:
    """Detect if a PDF page has broken/unmapped fonts (CID encoding)."""
    try:
        words = page.get_text("words")
        if not words:
            return True
        # Sample some words and check if they contain only low codepoints (CID garbage)
        sample = words[:20]
        garbage_count = 0
        for w in sample:
            text = str(w[4]) if len(w) > 4 else ""
            if not text:
                continue
            # Check if most chars are control chars or CID-mapped garbage
            low_chars = sum(1 for c in text if ord(c) < 32 or c in '!"#$%&\'()*+')
            if low_chars > len(text) * 0.3:
                garbage_count += 1
        return garbage_count > len(sample) * 0.4
    except Exception:
        return True


def easyocr_page_words(page: fitz.Page, page_number: int, lang: str, dpi: int = 200) -> List[Dict[str, Any]]:
    """Extract words from a page using EasyOCR."""
    if easyocr is None or np is None:
        raise RuntimeError("EasyOCR non installé. pip install easyocr numpy")

    pix = page.get_pixmap(dpi=dpi, alpha=False)
    img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)

    reader = get_easyocr_reader(lang)
    results = reader.readtext(img_np, text_threshold=0.3, low_text=0.3, min_size=5)

    # Scale factor: OCR coords are in pixels at given DPI, PDF coords are at 72 DPI
    scale = 72.0 / dpi

    words: List[Dict[str, Any]] = []
    for i, (bbox, text, conf) in enumerate(results):
        text = safe_excel_text(text)
        if not text or conf < 0.15:
            continue
        x0 = float(bbox[0][0]) * scale
        y0 = float(bbox[0][1]) * scale
        x1 = float(bbox[2][0]) * scale
        y1 = float(bbox[2][1]) * scale
        if x1 <= x0 or y1 <= y0:
            continue
        words.append({
            "page": page_number,
            "text": text,
            "x0": x0,
            "x1": x1,
            "top": y0,
            "bottom": y1,
            "xc": (x0 + x1) / 2.0,
            "yc": (y0 + y1) / 2.0,
            "block_num": 0,
            "par_num": 0,
            "line_num": i,
            "word_num": i,
        })
    return words


def render_page_to_image(page: fitz.Page, dpi: int = 200) -> Image.Image:
    scale = dpi / 72.0
    matrix = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=matrix, alpha=False, colorspace=fitz.csGRAY)
    img = Image.frombytes("L", [pix.width, pix.height], pix.samples)
    img = img.point(lambda p: 255 if p > 210 else 0)
    return img

def tesseract_available() -> bool:
    if pytesseract is None:
        return False
    try:
        return bool(pytesseract.get_tesseract_version())
    except Exception:
        return False

def page_words_from_pdf_text(page: fitz.Page, page_number: int) -> List[Dict[str, Any]]:
    words = page.get_text("words")
    out: List[Dict[str, Any]] = []
    for idx, item in enumerate(words):
        if len(item) < 5:
            continue
        x0, y0, x1, y1, text = item[:5]
        text = safe_excel_text(text)
        if not text:
            continue
        block_num = item[5] if len(item) > 5 else 0
        line_num = item[6] if len(item) > 6 else idx
        word_num = item[7] if len(item) > 7 else idx
        out.append({
            "page": page_number,
            "text": text,
            "x0": float(x0),
            "x1": float(x1),
            "top": float(y0),
            "bottom": float(y1),
            "xc": (float(x0) + float(x1)) / 2.0,
            "yc": (float(y0) + float(y1)) / 2.0,
            "block_num": int(block_num),
            "par_num": 0,
            "line_num": int(line_num),
            "word_num": int(word_num),
        })
    return out

def ocr_page_words(page: fitz.Page, page_number: int, lang: str) -> List[Dict[str, Any]]:
    if pytesseract is None:
        raise RuntimeError(
            "Le module Python 'pytesseract' n'est pas installé. "
            "Installe-le avec: py -m pip install pytesseract"
        )
    if not tesseract_available():
        raise RuntimeError(
            "Tesseract OCR n'est pas disponible. Vérifie l'installation Windows de Tesseract "
            "ou ajuste les chemins TESSERACT_EXE / TESSDATA_DIR dans le script."
        )

    img = render_page_to_image(page, dpi=220)
    data = pytesseract.image_to_data(
        img,
        lang=lang,
        output_type=pytesseract.Output.DICT,
        config="--oem 3 --psm 6 preserve_interword_spaces=1",
    )

    words: List[Dict[str, Any]] = []
    for i in range(len(data["text"])):
        text = safe_excel_text(data["text"][i])
        if not text:
            continue
        try:
            conf = float(data["conf"][i])
        except Exception:
            conf = -1
        if conf < 0:
            continue

        left = float(data["left"][i])
        top = float(data["top"][i])
        width = float(data["width"][i])
        height = float(data["height"][i])
        if width <= 0 or height <= 0:
            continue

        words.append({
            "page": page_number,
            "text": text,
            "x0": left,
            "x1": left + width,
            "top": top,
            "bottom": top + height,
            "xc": left + (width / 2.0),
            "yc": top + (height / 2.0),
            "block_num": data["block_num"][i],
            "par_num": data["par_num"][i],
            "line_num": data["line_num"][i],
            "word_num": i,
        })
    return words

def group_words_to_lines(words: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not words:
        return []
    words = sorted(words, key=lambda w: (w["page"], w["top"], w["x0"]))

    lines: List[Dict[str, Any]] = []
    current: List[Dict[str, Any]] = []
    tolerance = 8.0  # Increased for OCR output

    for w in words:
        if not current:
            current = [w]
            continue

        same_page = w["page"] == current[0]["page"]
        current_y = statistics.median([cw["yc"] for cw in current])
        if same_page and abs(w["yc"] - current_y) <= tolerance:
            current.append(w)
        else:
            current_sorted = sorted(current, key=lambda x: x["x0"])
            lines.append({
                "page": current_sorted[0]["page"],
                "top": min(x["top"] for x in current_sorted),
                "bottom": max(x["bottom"] for x in current_sorted),
                "words": current_sorted,
                "text": safe_excel_text(" ".join(x["text"] for x in current_sorted)),
            })
            current = [w]

    if current:
        current_sorted = sorted(current, key=lambda x: x["x0"])
        lines.append({
            "page": current_sorted[0]["page"],
            "top": min(x["top"] for x in current_sorted),
            "bottom": max(x["bottom"] for x in current_sorted),
            "words": current_sorted,
            "text": safe_excel_text(" ".join(x["text"] for x in current_sorted)),
        })

    return lines

def detect_column_positions(lines: List[Dict[str, Any]]) -> Dict[str, float]:
    header_candidates: List[Tuple[str, float]] = []
    for line in lines[:40]:
        txt = line["text"].lower()
        for w in line["words"]:
            wt = w["text"].lower().strip()
            # Match column headers more flexibly (handles OCR word boundaries)
            if "date" in wt and ("opér" in wt or "oper" in wt):
                header_candidates.append(("date_op", w["x0"]))
            elif wt == "date" and ("opér" in txt or "oper" in txt):
                header_candidates.append(("date_op", w["x0"]))
            if "libell" in wt:
                header_candidates.append(("libelle", w["x0"]))
            if "valeur" in wt:
                header_candidates.append(("date_valeur", w["x0"]))
            elif wt == "date" and "valeur" in txt and "opér" not in txt:
                header_candidates.append(("date_valeur", w["x0"]))
            if "débit" in wt or "debit" in wt:
                header_candidates.append(("debit", w["x0"]))
            if ("crédit" in wt or "credit" in wt) and "créditeur" not in wt and "crediteur" not in wt:
                header_candidates.append(("credit", w["x0"]))
            if "solde" in wt and "ancien" not in txt and "nouveau" not in txt:
                header_candidates.append(("solde", w["x0"]))

    # Also scan raw words from the first lines (bypass line grouping issues)
    if not header_candidates:
        for line in lines[:40]:
            for w in line["words"]:
                wt = w["text"].lower().strip()
                if wt in ("débit", "debit"):
                    header_candidates.append(("debit", w["x0"]))
                elif wt in ("crédit", "credit"):
                    header_candidates.append(("credit", w["x0"]))
                elif wt == "solde":
                    header_candidates.append(("solde", w["x0"]))
                elif "libell" in wt:
                    header_candidates.append(("libelle", w["x0"]))

    buckets: Dict[str, List[float]] = defaultdict(list)
    for name, pos in header_candidates:
        buckets[name].append(pos)

    # Defaults calibrated for OCR with 72-DPI scaled coordinates
    # These work for typical Caisse d'Epargne / bank statement layouts
    defaults = {
        "date_op": 35.0,
        "libelle": 90.0,
        "date_valeur": 240.0,
        "debit": 310.0,
        "credit": 355.0,
        "solde": 400.0,
    }

    result: Dict[str, float] = {}
    for key, default in defaults.items():
        vals = buckets.get(key)
        result[key] = statistics.median(vals) if vals else default

    # Sanity check: positions must be in order date_op < libelle < date_valeur < debit < credit < solde
    ordered = ["date_op", "libelle", "date_valeur", "debit", "credit", "solde"]
    for i in range(len(ordered) - 1):
        if result[ordered[i]] >= result[ordered[i + 1]]:
            # Position out of order — reset to defaults
            print(f"  ⚠ Column {ordered[i]}={result[ordered[i]]:.1f} >= {ordered[i+1]}={result[ordered[i+1]]:.1f}, using defaults", flush=True)
            return defaults

    print(f"  Column positions: {', '.join(f'{k}={v:.1f}' + ('*' if buckets.get(k) else '') for k, v in result.items())}", flush=True)
    return result

def split_by_boundaries(words: List[Dict[str, Any]], pos: Dict[str, float]) -> Dict[str, str]:
    bounds = [
        ("date_op", pos["date_op"], pos["libelle"]),
        ("libelle", pos["libelle"], pos["date_valeur"]),
        ("date_valeur", pos["date_valeur"], pos["debit"]),
        ("debit", pos["debit"], pos["credit"]),
        ("credit", pos["credit"], pos["solde"]),
        ("solde", pos["solde"], 10_000.0),
    ]

    cells = {k: "" for k, _, _ in bounds}
    for w in words:
        x = w["xc"]
        matched = False
        for key, left, right in bounds:
            if left - 8 <= x < right - 8:
                cells[key] = safe_excel_text((cells[key] + " " + w["text"]).strip())
                matched = True
                break
        if not matched:
            # Word fell outside all boundaries — try closest column
            min_dist = float('inf')
            best_key = "libelle"
            for key, left, right in bounds:
                mid = (left + right) / 2
                dist = abs(x - mid)
                if dist < min_dist:
                    min_dist = dist
                    best_key = key
            cells[best_key] = safe_excel_text((cells[best_key] + " " + w["text"]).strip())
    return cells

def normalize_row(cells: Dict[str, str]) -> Dict[str, str]:
    for key in list(cells.keys()):
        cells[key] = safe_excel_text(cells.get(key, ""))

    # Date opération contient aussi le libellé (ex: "02/01/2023 COMMISSIONS")
    if cells["date_op"]:
        m = DATE_RE.match(cells["date_op"])
        if m and len(cells["date_op"]) > len(m.group(0)) + 1:
            extra = cells["date_op"][len(m.group(0)):].strip(" _-|';:")
            if extra and not DATE_RE.match(extra):
                cells["date_op"] = m.group(0)
                cells["libelle"] = safe_excel_text((extra + " " + cells["libelle"]).strip()) if cells["libelle"] else extra

    # Date opération qui aurait glissé dans le libellé
    if not cells["date_op"]:
        m = DATE_RE.match(cells["libelle"])
        if m:
            cells["date_op"] = m.group(0)
            cells["libelle"] = safe_excel_text(cells["libelle"][len(m.group(0)):].lstrip(" _-|';:"))

    # Date de valeur glissée dans libellé
    if not cells["date_valeur"]:
        m = DATE_RE.search(cells["libelle"])
        if m and cells["libelle"].startswith(m.group(0)):
            cells["date_valeur"] = m.group(0)
            cells["libelle"] = safe_excel_text(cells["libelle"][len(m.group(0)):].lstrip(" _-|';:"))

    # Montant collé dans date_valeur
    if cells["date_valeur"] and not DATE_RE.search(cells["date_valeur"]):
        m = AMOUNT_RE.search(cells["date_valeur"])
        if m:
            if not cells["debit"] and not cells["credit"]:
                cells["debit"] = m.group(0)
            cells["date_valeur"] = ""

    return cells

def looks_like_header(text: str) -> bool:
    t = text.lower()
    if ("date" in t and "libell" in t) or ("débit" in t or "debit" in t) and ("solde" in t):
        return True
    # Filter "Ancien Solde" / "Nouveau Solde" info lines
    # Note: "Ancien Solde" lines are kept as data rows (initial balance)
    # Filter page footers (URLs, page numbers like "1/54")
    if "https://" in t or "http://" in t:
        return True
    if re.match(r"^\d+/\d+$", t.strip()):
        return True
    return False

def extract_rows(pdf_bytes: bytes, lang: str, job_id: str, ocr_mode: str, exclude_zones: list = None, headers_data: list = None) -> Dict[str, Any]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total_pages = len(doc)
    if total_pages == 0:
        doc.close()
        return {"rows": []}

    all_lines: List[Dict[str, Any]] = []
    used_mode = None

    # Auto-detect broken fonts on first page
    force_ocr = False
    if ocr_mode == "auto":
        first_page = doc[0]
        if has_broken_fonts(first_page):
            print("  ⚠ Broken fonts detected → forcing EasyOCR mode")
            force_ocr = True

    for page_number, page in enumerate(doc, start=1):
        pct = int(((page_number - 1) / total_pages) * 75)
        # Check if job was cancelled
        with JOBS_LOCK:
            if JOBS.get(job_id, {}).get("status") == "error":
                doc.close()
                return {"rows": [], "used_mode": "cancelled"}

        set_job(job_id, progress=pct, message=f"Analyse page {page_number}/{total_pages} ({'OCR' if force_ocr or ocr_mode == 'ocr_only' else 'texte'})...")

        words: List[Dict[str, Any]] = []

        if force_ocr or ocr_mode == "ocr_only":
            # Use EasyOCR (preferred) or fallback to pytesseract
            try:
                words = easyocr_page_words(page, page_number, lang)
                used_mode = "easyocr"
            except RuntimeError:
                words = ocr_page_words(page, page_number, lang)
                used_mode = "tesseract"
        elif ocr_mode in ("auto", "text_only"):
            words = page_words_from_pdf_text(page, page_number)
            # If too few words, try OCR
            if ocr_mode == "auto" and len(words) < 15:
                try:
                    words = easyocr_page_words(page, page_number, lang)
                    used_mode = "easyocr"
                except RuntimeError:
                    try:
                        words = ocr_page_words(page, page_number, lang)
                        used_mode = "tesseract"
                    except RuntimeError:
                        pass  # keep whatever text extraction found
            elif used_mode is None:
                used_mode = "text"

        # Filter out words in exclude zones
        if exclude_zones and words:
            page_rect = page.rect  # PDF page dimensions
            pw, ph = page_rect.width, page_rect.height
            filtered = []
            for w in words:
                excluded = False
                for z in exclude_zones:
                    z_from = z.get("fromPage", z.get("p", 0))
                    z_to = z.get("toPage", z_from)
                    z_all = z.get("a", False)
                    if not z_all and not (z_from <= page_number <= z_to):
                        continue
                    # Zone coords are 0-1 relative to image; word coords are in PDF points
                    zx0 = z["x"] * pw
                    zy0 = z["y"] * ph
                    zx1 = (z["x"] + z["w"]) * pw
                    zy1 = (z["y"] + z["h"]) * ph
                    if zx0 <= w["xc"] <= zx1 and zy0 <= w["yc"] <= zy1:
                        excluded = True
                        break
                if not excluded:
                    filtered.append(w)
            words = filtered

        all_lines.extend(group_words_to_lines(words))

    doc.close()

    set_job(job_id, progress=78, message="Détection des colonnes...")
    # Sort headers by page then by lineY
    sorted_headers = []
    if headers_data:
        sorted_headers = sorted(headers_data, key=lambda h: (h.get("page", 1), h.get("lineY", 0)))
        for sh in sorted_headers:
            print(f"  Header on page {sh.get('page',1)}: {sh.get('cols',[])} ({len(sh.get('positions',[]))} positions)", flush=True)

    if sorted_headers and sorted_headers[0].get("cols") and sorted_headers[0].get("positions"):
        # Use first header as default
        generic_mode = True
        col_names = sorted_headers[0]["cols"]
        col_pos = sorted_headers[0]["positions"]
    else:
        # Auto-detect from lines: find the best header line
        print("  No header_cols provided, auto-detecting from lines...", flush=True)
        best_line = None
        best_score = 0
        for line in all_lines[:40]:
            words = line["words"]
            if len(words) < 3:
                continue
            has_date_or_amount = any(DATE_RE.search(w["text"]) or AMOUNT_RE.search(w["text"]) for w in words)
            if has_date_or_amount:
                continue
            decimal_count = sum(1 for w in words if re.match(r"^[\d][\d .,]*[\d]$", w["text"].strip()) and len(w["text"].strip()) > 3)
            if decimal_count >= len(words) * 0.5:
                continue
            ltxt = " ".join(w["text"] for w in words).lower()
            if "http" in ltxt:
                continue
            avg_len = sum(len(w["text"]) for w in words) / len(words)
            if avg_len > 20:
                continue
            texts = [w["text"].lower() for w in words]
            if len(set(texts)) < len(texts) * 0.7:
                continue
            xs = [w["xc"] for w in words]
            ys = [w["yc"] for w in words]
            if len(ys) > 1 and (max(ys) - min(ys)) > 5:
                continue
            score = len(words) * (max(xs) - min(xs))
            if score > best_score:
                best_score = score
                best_line = line

        # Strategy 2: find first data row and take line above
        if not best_line and len(all_lines) >= 2:
            for idx in range(1, min(40, len(all_lines))):
                words_l = all_lines[idx]["words"]
                if len(words_l) < 3: continue
                num_count = sum(1 for w in words_l if re.match(r"^\d+[.,]\d+$", w["text"].strip().replace(" ","")))
                if num_count >= len(words_l) * 0.5:
                    best_line = all_lines[idx - 1]
                    break

        if best_line:
            generic_mode = True
            col_names = [w["text"] for w in sorted(best_line["words"], key=lambda w: w["x0"])]
            col_pos = [w["x0"] for w in sorted(best_line["words"], key=lambda w: w["x0"])]
            print(f"  Auto-detected {len(col_names)} columns: {col_names}", flush=True)
        else:
            # Last resort fallback
            generic_mode = False
            pos = detect_column_positions(all_lines)
            col_names = None
            col_pos = None
            print("  Could not auto-detect, using legacy fallback", flush=True)

    rows: List[Dict[str, Any]] = []
    current: Optional[Dict[str, str]] = None
    total_lines = max(len(all_lines), 1)

    for i, line in enumerate(all_lines, start=1):
        if i % 20 == 0 or i == total_lines:
            pct = 78 + int((i / total_lines) * 18)
            set_job(job_id, progress=min(pct, 96), message=f"Construction des lignes {i}/{total_lines}...")

        if looks_like_header(line["text"]):
            continue

        if generic_mode:
            # Find applicable header for this line
            cur_col_names = col_names
            cur_col_pos = col_pos
            hdr_idx = 0
            if len(sorted_headers) > 1:
                line_page = line["page"]
                for si, sh in enumerate(reversed(sorted_headers)):
                    sh_page = sh.get("page", 1)
                    if line_page >= sh_page:
                        if sh.get("cols") and sh.get("positions"):
                            cur_col_names = sh["cols"]
                            cur_col_pos = sh["positions"]
                            hdr_idx = len(sorted_headers) - 1 - si
                        break
            cells_g = {}
            n = len(cur_col_names)
            for w in line["words"]:
                x = w["xc"]
                best_col = cur_col_names[-1]
                for ci in range(n):
                    left = (cur_col_pos[ci-1]+cur_col_pos[ci])/2 if ci>0 else 0
                    right = (cur_col_pos[ci]+cur_col_pos[ci+1])/2 if ci+1<n else float('inf')
                    if left <= x < right:
                        best_col = cur_col_names[ci]
                        break
                cells_g[best_col] = safe_excel_text((cells_g.get(best_col,"") + " " + w["text"]).strip())
            # Skip if empty
            if not any(v.strip() for v in cells_g.values()):
                continue
            # Skip repeated headers
            if all(cells_g.get(cn,"").strip() == cn for cn in cur_col_names if cells_g.get(cn,"").strip()):
                continue
            # Smart alignment: if first column has non-date text and second column is empty,
            # shift content from first to second (e.g., 'Ancien Solde au' misplaced in Date column)
            if len(cur_col_names) >= 2:
                first_val = cells_g.get(cur_col_names[0], "").strip()
                second_val = cells_g.get(cur_col_names[1], "").strip()
                if first_val and not second_val and not DATE_RE.search(first_val) and not AMOUNT_RE.search(first_val):
                    cells_g[cur_col_names[1]] = first_val
                    cells_g[cur_col_names[0]] = ""
            rows.append({"page": line["page"], "_hdr": hdr_idx, **cells_g})
            continue
        cells = normalize_row(split_by_boundaries(line["words"], pos))

        line_has_date_op = bool(DATE_RE.fullmatch(cells["date_op"]))
        line_has_any_amount = any(cells[k] for k in ("debit", "credit", "solde"))

        if line_has_date_op:
            if current and any(current.values()):
                rows.append({"page": line["page"], **current})
            current = {
                "date_operation": cells["date_op"],
                "libelle": cells["libelle"],
                "date_valeur": cells["date_valeur"],
                "debit": cells["debit"],
                "credit": cells["credit"],
                "solde": cells["solde"],
            }
            continue

        if current is None:
            # ligne isolée type "Ancien Solde au"
            if any(cells.values()):
                rows.append({
                    "page": line["page"],
                    "date_operation": cells["date_op"],
                    "libelle": cells["libelle"],
                    "date_valeur": cells["date_valeur"],
                    "debit": cells["debit"],
                    "credit": cells["credit"],
                    "solde": cells["solde"],
                })
            continue

        # Ligne de continuation du libellé
        if cells["libelle"] and not line_has_any_amount and not cells["date_valeur"]:
            current["libelle"] = safe_excel_text((current["libelle"] + " " + cells["libelle"]).strip())
            continue

        # Ligne qui complète la transaction précédente
        if cells["libelle"] and not current["libelle"]:
            current["libelle"] = cells["libelle"]
        elif cells["libelle"] and not line_has_any_amount and cells["date_valeur"]:
            current["libelle"] = safe_excel_text((current["libelle"] + " " + cells["libelle"]).strip())

        for src, dst in [
            ("date_valeur", "date_valeur"),
            ("debit", "debit"),
            ("credit", "credit"),
            ("solde", "solde"),
        ]:
            if cells[src] and not current[dst]:
                current[dst] = cells[src]

    if current and any(current.values()):
        rows.append({"page": all_lines[-1]["page"] if all_lines else 1, **current})

    result = {"rows": rows, "used_mode": used_mode or ocr_mode}
    if generic_mode:
        result["col_names"] = col_names
        result["all_headers"] = sorted_headers
    return result

def build_workbook(data: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    all_headers = data.get("all_headers", [])
    generic_cols = data.get("col_names")

    # Multiple headers → one sheet per header
    if all_headers and len(all_headers) > 1 and generic_cols:
        for hi, hdr in enumerate(all_headers):
            hdr_cols = hdr.get("cols", generic_cols)
            if hi == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet()
            ws.title = f"Tableau {hi+1}"
            ws.append(["Page"] + hdr_cols)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in data["rows"]:
                if row.get("_hdr", 0) == hi:
                    ws.append([row.get("page", "")] + [safe_excel_text(row.get(col, "")) for col in hdr_cols])
        return wb

    # Single header or legacy mode
    ws = wb.active
    ws.title = "Donnees"
    if generic_cols:
        headers = ["Page"] + generic_cols
    else:
        headers = ["Page", "Date opération", "Libellé", "Date de valeur", "Débit", "Crédit", "Solde"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in data["rows"]:
        if generic_cols:
            ws.append([row.get("page", "")] + [safe_excel_text(row.get(col, "")) for col in generic_cols])
        else:
            ws.append([
                row.get("page", ""),
                safe_excel_text(row.get("date_operation", "")),
                safe_excel_text(row.get("libelle", "")),
                safe_excel_text(row.get("date_valeur", "")),
                safe_excel_text(row.get("debit", "")),
                safe_excel_text(row.get("credit", "")),
                safe_excel_text(row.get("solde", "")),
            ])

    widths = {"A": 8, "B": 16, "C": 52, "D": 16, "E": 14, "F": 14, "G": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(ws.max_row, 2)}"

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 8):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    return wb

def process_job(job_id: str, pdf_bytes: bytes, filename: str, lang: str, ocr_mode: str, exclude_zones: list = None, headers_data: list = None) -> None:
    try:
        set_job(job_id, status="running", progress=2, message="Ouverture du PDF...")
        data = extract_rows(pdf_bytes, lang=lang, job_id=job_id, ocr_mode=ocr_mode, exclude_zones=exclude_zones or [], headers_data=headers_data)

        set_job(job_id, progress=97, message=f"Création du fichier XLSX... (mode utilisé: {data.get('used_mode', ocr_mode)})")
        wb = build_workbook(data)

        out_name = os.path.splitext(filename)[0] + "_xlsx.xlsx"
        out_path = os.path.join(OUTPUT_DIR, f"{job_id}_{out_name}")

        with io.BytesIO() as output:
            wb.save(output)
            with open(out_path, "wb") as f:
                f.write(output.getvalue())

        set_job(
            job_id,
            status="done",
            progress=100,
            message=f"Conversion terminée. Mode utilisé: {data.get('used_mode', ocr_mode)}",
            output_path=out_path,
            download_name=out_name,
        )
    except Exception as e:
        set_job(job_id, status="error", progress=100, error=str(e), message="Échec de conversion.")

# Uploaded files store: file_id -> {path, filename, total_pages}
UPLOADS: Dict[str, Dict[str, Any]] = {}

@app.get("/")
def index():
    return render_template_string(HTML_FORM)

@app.get("/logo.png")
def logo():
    return send_file(os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png"), mimetype="image/png")

@app.post("/upload")
def upload():
    """Pre-upload PDF for preview. Returns file_id + total_pages."""
    if "pdf" not in request.files:
        return jsonify({"error": "Aucun fichier recu."}), 400
    file = request.files["pdf"]
    if not file or not file.filename:
        return jsonify({"error": "Aucun fichier selectionne."}), 400
    filename = secure_filename(file.filename)
    if not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Le fichier doit etre un PDF."}), 400
    pdf_bytes = file.read()
    if not pdf_bytes:
        return jsonify({"error": "Fichier vide."}), 400

    file_id = uuid.uuid4().hex
    pdf_path = os.path.join(OUTPUT_DIR, f"{file_id}.pdf")
    with open(pdf_path, "wb") as f:
        f.write(pdf_bytes)

    # Count pages
    try:
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        doc.close()
    except Exception as e:
        return jsonify({"error": f"PDF invalide: {e}"}), 400

    UPLOADS[file_id] = {"path": pdf_path, "filename": filename, "total_pages": total_pages}
    return jsonify({"file_id": file_id, "total_pages": total_pages})

@app.get("/preview/<file_id>/<int:page>")
def preview(file_id: str, page: int):
    """Render a PDF page as PNG for preview."""
    info = UPLOADS.get(file_id)
    if not info:
        abort(404, "Fichier introuvable.")
    if page < 1 or page > info["total_pages"]:
        abort(400, "Page invalide.")

    doc = fitz.open(info["path"])
    pg = doc[page - 1]
    pix = pg.get_pixmap(dpi=150, alpha=False)
    doc.close()

    img_bytes = pix.tobytes("png")
    return Response(img_bytes, mimetype="image/png")


@app.get("/auto_detect_header/<file_id>/<int:page>")
def auto_detect_header(file_id: str, page: int):
    """Auto-detect the table header line on a page via OCR."""
    info = UPLOADS.get(file_id)
    if not info:
        return jsonify({"columns":[],"positions":[],"y":0})
    try:
        doc = fitz.open(info["path"])
        pg = doc[page - 1]
        page_h = pg.rect.height
        page_w = pg.rect.width
        pix = pg.get_pixmap(dpi=200, alpha=False)
        img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
        reader = get_easyocr_reader("fr")
        results = reader.readtext(img_np, text_threshold=0.3, low_text=0.3, min_size=5)
        scale = 72.0 / 200
        doc.close()
        # Build words with positions
        words = []
        for bbox, text, conf in results:
            if not text.strip() or conf < 0.15: continue
            x0, x1 = bbox[0][0]*scale, bbox[2][0]*scale
            yc = ((bbox[0][1]+bbox[2][1])/2)*scale
            words.append({"text":text,"x0":x0,"x1":x1,"xc":(x0+x1)/2,"yc":yc})
        words.sort(key=lambda w:(w["yc"],w["x0"]))
        # Remove stray "%" or "8" (OCR reads % as 8) that follow a number
        cleaned = []
        for i, w in enumerate(words):
            txt = w["text"].strip()
            if txt in ('8', '%', ';') and i > 0 and abs(w["yc"]-words[i-1]["yc"]) <= 8:
                prev = words[i-1]["text"].strip().replace(' ','')
                if re.match(r'^\d+', prev):
                    continue  # skip stray % symbol
            cleaned.append(w)
        words = cleaned
        # Also clean "8" suffix from merged words like "90 8" -> "90"
        for w in words:
            w["text"] = re.sub(r'\s+8$', '', w["text"])
            w["text"] = re.sub(r'\s+;$', '', w["text"])
        # Group into lines
        lines = []
        cur = []
        for w in words:
            if not cur or abs(w["yc"]-statistics.median(cw["yc"] for cw in cur))<=8:
                cur.append(w)
            else:
                lines.append(sorted(cur, key=lambda x:x["x0"]))
                cur = [w]
        if cur: lines.append(sorted(cur, key=lambda x:x["x0"]))
        # Score lines: best header = ALL words are text (no numbers/dates), max words*spread
        best = None
        best_score = 0
        best_y = 0
        for line in lines:
            if len(line)<3: continue
            has_date_or_amount = any(DATE_RE.search(w["text"]) or AMOUNT_RE.search(w["text"]) for w in line)
            if has_date_or_amount: continue
            # Reject lines where most words are decimal numbers (data rows)
            decimal_count = sum(1 for w in line if re.match(r"^[\d][\d .,]*[\d]$", w["text"].strip()) and len(w["text"].strip()) > 3)
            if decimal_count >= len(line) * 0.5: continue
            # Also reject lines where most words are purely numeric (even short)
            num_count = sum(1 for w in line if re.match(r'^[\d.,]+$', w["text"].strip()))
            if num_count >= len(line) * 0.6: continue
            ltxt = " ".join(w["text"] for w in line).lower()
            if "http" in ltxt: continue
            # Header words are typically short (< 20 chars avg)
            avg_len = sum(len(w["text"]) for w in line) / len(line)
            if avg_len > 20: continue
            # Header words should be unique (not repeating like "VIR SEPA ...")
            texts = [w["text"].lower() for w in line]
            if len(set(texts)) < len(texts) * 0.7: continue
            xs=[w["xc"] for w in line]
            # Prefer lines closer to top of page (headers are usually near top)
            avg_y = sum(w["yc"] for w in line) / len(line)
            y_bonus = max(0, 1.0 - avg_y / page_h) 
            score=len(line)*(max(xs)-min(xs)) * (1 + y_bonus)
            if score>best_score:
                best_score=score
                best=line
                best_y=statistics.median(w["yc"] for w in line)
        # Strategy 2: if no header found, find first data row and take the line above
        if not best and len(lines) >= 2:
            for idx in range(1, len(lines)):
                words_l = lines[idx]
                if len(words_l) < 3: continue
                num_count = sum(1 for w in words_l if re.match(r"^\d+[.,]\d+$", w["text"].strip().replace(" ","")))
                if num_count >= len(words_l) * 0.5:
                    # This is likely a data row — take the line above as header
                    best = lines[idx - 1]
                    best_y = sum(w["yc"] for w in best) / len(best)
                    break

        if best:
            return jsonify({"columns":[w["text"] for w in best],"positions":[w["x0"] for w in best],"y":best_y/page_h,"page_width":page_w})
    except Exception as e:
        print(f"  Auto-detect error: {e}", flush=True)
    return jsonify({"columns":[],"positions":[],"y":0})

@app.post("/detect_header_at/<file_id>/<int:page>")
def detect_header_at(file_id: str, page: int):
    """Detect columns at a specific Y position (user click)."""
    info = UPLOADS.get(file_id)
    if not info:
        return jsonify({"columns":[],"positions":[]})
    body = request.get_json()
    rel_y = body.get("y", 0.5)
    try:
        doc = fitz.open(info["path"])
        pg = doc[page - 1]
        page_h = pg.rect.height
        page_w = pg.rect.width
        target_y = rel_y * page_h
        tol = page_h * 0.01
        pix = pg.get_pixmap(dpi=200, alpha=False)
        img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
        reader = get_easyocr_reader("fr")
        results = reader.readtext(img_np, text_threshold=0.3, low_text=0.3, min_size=5)
        scale = 72.0 / 200
        doc.close()
        hw = []
        for bbox, text, conf in results:
            if not text.strip() or conf < 0.15: continue
            yc = ((bbox[0][1]+bbox[2][1])/2)*scale
            if abs(yc - target_y) <= tol:
                x0 = bbox[0][0]*scale
                x1 = bbox[2][0]*scale
                hw.append({"text":text,"x0":x0,"x1":x1,"yc":yc})
        hw.sort(key=lambda w:w["x0"])
        # Remove stray "%" or "8" (OCR reads % as 8) that follow a number
        cleaned_hw = []
        for i, w in enumerate(hw):
            txt = w["text"].strip()
            if txt in ('8', '%', ';') and i > 0:
                prev = hw[i-1]["text"].strip().replace(' ','')
                if re.match(r'^\d+', prev):
                    continue
            cleaned_hw.append(w)
        hw = cleaned_hw
        for w in hw:
            w["text"] = re.sub(r'\s+8$', '', w["text"])
            w["text"] = re.sub(r'\s+;$', '', w["text"])
        # Reject if data row (dates, amounts, or mostly numeric)
        if hw:
            num_count = sum(1 for w in hw if re.match(r'^[\d.,\s]+$', w["text"].strip()) or w["text"].strip() in ('.', ',', ';', ':', '0'))
            date_count = sum(1 for w in hw if DATE_RE.search(w["text"]))
            amount_count = sum(1 for w in hw if AMOUNT_RE.search(w["text"]))
            if num_count >= len(hw) * 0.5 or date_count >= 2 or amount_count >= 2:
                return jsonify({"columns":[],"positions":[],"page_width":page_w,"error":"data_row"})
        return jsonify({"columns":[w["text"] for w in hw],"positions":[w["x0"] for w in hw],"page_width":page_w})
    except Exception as e:
        return jsonify({"columns":[],"positions":[],"page_width":0})

@app.post("/start")
def start():
    """Start conversion. Accepts JSON (new) or form data (legacy)."""
    # New JSON mode (from new interface)
    if request.is_json:
        body = request.get_json()
        file_id = body.get("file_id")
        lang = body.get("lang", "fr")
        exclude_zones = body.get("exclude_zones", [])
        headers_data = body.get("headers", [])
        # Backward compat: single header
        header_cols = body.get("header_cols", [])
        header_positions = body.get("header_positions", [])
        if not headers_data and header_cols:
            headers_data = [{"page": 1, "cols": header_cols, "positions": header_positions}]

        info = UPLOADS.get(file_id)
        if not info:
            return jsonify({"error": "Fichier introuvable. Veuillez re-uploader."}), 400

        with open(info["path"], "rb") as f:
            pdf_bytes = f.read()
        filename = info["filename"]
        ocr_mode = "ocr_only"  # Always OCR for reliability
    else:
        # Legacy form mode
        if "pdf" not in request.files:
            return jsonify({"error": "Aucun fichier recu."}), 400
        file = request.files["pdf"]
        if not file or not file.filename:
            return jsonify({"error": "Aucun fichier selectionne."}), 400
        filename = secure_filename(file.filename)
        if not filename.lower().endswith(".pdf"):
            return jsonify({"error": "Le fichier doit etre un PDF."}), 400
        lang = (request.form.get("lang", "fr") or "fr").strip()
        ocr_mode = "auto"
        pdf_bytes = file.read()
        exclude_zones = []
        if not pdf_bytes:
            return jsonify({"error": "Fichier vide."}), 400

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "status": "queued",
            "progress": 0,
            "message": "En attente...",
            "error": None,
            "output_path": None,
            "download_name": None,
            "created_at": time.time(),
        }

    thread = threading.Thread(
        target=process_job,
        args=(job_id, pdf_bytes, filename, lang, ocr_mode),
        kwargs={"exclude_zones": exclude_zones, "headers_data": headers_data if request.is_json else None},
        daemon=True
    )
    thread.start()

    return jsonify({"job_id": job_id})

@app.get("/status/<job_id>")
def status(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)

    if not job:
        abort(404, "Job introuvable.")

    return jsonify({
        "status": job.get("status"),
        "progress": job.get("progress"),
        "message": job.get("message"),
        "error": job.get("error"),
    })

@app.post("/cancel/<job_id>")
def cancel(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        return jsonify({"ok": False}), 404
    set_job(job_id, status="error", error="Annule par l'utilisateur.", message="Conversion annulée.")
    return jsonify({"ok": True})

@app.get("/download/<job_id>")
def download(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)

    if not job:
        abort(404, "Job introuvable.")

    if job.get("status") != "done" or not job.get("output_path"):
        abort(400, "Fichier non prêt.")

    output_path = job["output_path"]
    download_name = job.get("download_name") or os.path.basename(output_path)

    if not os.path.exists(output_path):
        abort(404, "Fichier de sortie introuvable.")

    return send_file(
        output_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name
    )

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False, use_reloader=False, threaded=True)
